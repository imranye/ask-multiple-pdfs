import os
import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings, HuggingFaceInstructEmbeddings
from langchain.vectorstores import FAISS
from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from htmlTemplates import css, bot_template, user_template
from langchain.llms import HuggingFaceHub

load_dotenv()  # load environment variables from .env file

def get_doc_text(doc_files):
    text = ""
    for doc in doc_files:
        doc_reader = Document(doc)
        for para in doc_reader.paragraphs:
            text += para.text
    return text

def get_pdf_text(pdf_docs):
    text = ""
    for pdf in pdf_docs:
        pdf_reader = PdfReader(pdf)
        for page in pdf_reader.pages:
            text += page.extract_text()
    return text

def get_excel_text(excel_files):
    text = ""
    for excel in excel_files:
        wb = load_workbook(excel)
        for sheet in wb.sheetnames:
            df = pd.read_excel(excel, sheet_name=sheet)
            text += df.to_string()
    return text

def upload_files():
    doc_files = st.file_uploader("Upload DOCX", type=['docx', 'docx'], accept_multiple_files=True)
    excel_files = st.file_uploader("Upload Excel", type=['xlsx', 'xls'], accept_multiple_files=True)
    return doc_files, excel_files, pdf_docs

def get_text_chunks(text):
    text_splitter = CharacterTextSplitter(
        separator="\n",
        chunk_size=1000,
        chunk_overlap=200,
        length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks


def get_vectorstore(text_chunks):
    embeddings = OpenAIEmbeddings()
    # embeddings = HuggingFaceInstructEmbeddings(model_name="hkunlp/instructor-xl")
    vectorstore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
    return vectorstore


def get_conversation_chain(vectorstore):
    llm = ChatOpenAI()
    # llm = HuggingFaceHub(repo_id="google/flan-t5-xxl", model_kwargs={"temperature":0.5, "max_length":512})

    memory = ConversationBufferMemory(
        memory_key='chat_history', return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm,
        retriever=vectorstore.as_retriever(),
        memory=memory
    )
    return conversation_chain


def handle_userinput(user_question):
    modified_question = "Answer in extreme detail, make sure not to use jargon and to be concise but detailed and technical, assume you are talking to an extremely well educated attornery. Also remove any introductory text in your answer, answer should be technical and site names, dates, relevant cases, etc, always reference sources in line dont just summarize, the question is as follows:" + user_question
    response = st.session_state.conversation({'question': user_question})
    st.session_state.chat_history = response['chat_history']

    for i, message in enumerate(st.session_state.chat_history):
        if i % 2 == 0:
            st.write(user_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
        else:
            st.write(bot_template.replace(
                "{{MSG}}", message.content), unsafe_allow_html=True)
            if 'metadata' in message and 'source' in message.metadata:
                st.write(f"Source: {message.metadata['source']}")


def center_image(image_path):
    col1, col2, col3 = st.columns([1,2,1])

    with col1:
        st.write("")

    with col2:
        st.image(image_path, width=300)

    with col3:
        st.write("")

def main():
    load_dotenv()
    st.set_page_config(page_title="WNJ-GPT",
                       page_icon=":books:")
    st.markdown(css, unsafe_allow_html=True)
    st.write(css, unsafe_allow_html=True)
    center_image("./logo.png")
    st.markdown("<br>", unsafe_allow_html=True)

    if "conversation" not in st.session_state:
        st.session_state.conversation = None
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = None
    if "processed" not in st.session_state:
        st.session_state.processed = False

    st.markdown("<h2 style='text-align: center;'>WNJ-GPT</h2>", unsafe_allow_html=True)
    
    # if not st.session_state.processed:
    #     st.write("Please upload a document to get started.")

    st.markdown("<h4 style='text-align: center;'> Upload Documents for Review & Analysis </h4>", unsafe_allow_html=True)
    st.markdown("""
    <style>
    .stFileUploader > div:first-child {
        background-color: blue;
        color: white;
    }
    </style>
    """, unsafe_allow_html=True)
    docs = st.file_uploader(" ", accept_multiple_files=True, type=['pdf', 'docx', 'xlsx', 'xls'])
    if st.button("Process"):
        with st.spinner("Processing"):
            pdf_docs = [doc for doc in docs if doc.type == 'application/pdf']
            docx_docs = [doc for doc in docs if doc.type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document']
            excel_docs = [doc for doc in docs if doc.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']
            # get pdf text
            raw_text = get_pdf_text(pdf_docs)
            # get docx text
            raw_text += get_doc_text(docx_docs)
            # get excel text
            raw_text += get_excel_text(excel_docs)

            # get the text chunks
            text_chunks = get_text_chunks(raw_text)

            # create vector store
            vectorstore = get_vectorstore(text_chunks)

            # create conversation chain
            st.session_state.conversation = get_conversation_chain(
                vectorstore)
            
            # Set processed flag
            st.session_state.processed = True

    # Check processed flag and display confirmation message
    if st.session_state.processed:
        st.success("Documents uploaded successfully! ✅")
        user_question = st.text_input("Ask a question about your documents:")
        if st.button("Submit") or user_question:
            handle_userinput(user_question)

if __name__ == '__main__':
    main()





